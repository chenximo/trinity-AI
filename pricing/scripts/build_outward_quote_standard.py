#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build Trinity L3a outward quote sheet (数据宝形态).

- 目录价绝对价（入/出/缓存）+ 公开用量阶梯折（对外三档 $5k/$10k/$50k，入/出同折）
- 目录价含原缓存价列；阶梯列仍为入/出（缓存命中阶梯另议）
- 商务总册：读 10/11；回写 01_报价解析汇总（不改 10/11）
- 阶梯数字：定价方案-v0 公开定折矩阵（非 L3b 商务深折照抄）
- 本轮目录价：线上为主；刊例策略 §3.1 上浮两款覆盖（hy-mt2-plus / glm-5.2）

Sources:
  - L2 线上刊例: pricing/output/online/prices-api-text.json（主源，覆盖全量上架）
  - 可选对照: pricing/output/trinity-pricing-text.xlsx「刊例对比校验-生文」
  - 模型→成本族: pricing/output/商务洽谈折扣总表.xlsx（10 归属 + 11 推荐成本折）
  - 回写商务表: 01_报价解析汇总（报价依据）
"""
from __future__ import annotations

import json
import math
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_SCRIPTS = Path(__file__).resolve().parent
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))
from xlsx_wechat_compat import patch_xlsx_for_wechat  # noqa: E402


def solid_fill(rgb6: str) -> PatternFill:
    """不透明实心底色。须 FF 前缀：openpyxl 默认 00=透明，微信预览会整格无色。"""
    h = str(rgb6).strip().lstrip("#").upper()
    if len(h) == 8:
        h = "FF" + h[2:]
    elif len(h) == 6:
        h = "FF" + h
    return PatternFill(fill_type="solid", fgColor=h, bgColor=h)

ROOT = Path(__file__).resolve().parents[1]
SRC_ONLINE = ROOT / "output/online/prices-api-text.json"
SRC_ONLINE_IMAGE = ROOT / "output/online/prices-api-image.json"
SRC_ONLINE_VIDEO = ROOT / "output/online/prices-api-video.json"
SRC_PRICES = ROOT / "output/trinity-pricing-text.xlsx"
SRC_COMMERCIAL = ROOT / "output/商务洽谈折扣总表.xlsx"
ROUTES_TEXT = ROOT / "input/routes-20260809-text"
ROUTES_IMAGE = ROOT / "input/routes-20260809-image"
ROUTES_VIDEO = ROOT / "input/routes-20260809-video"
OUT = ROOT / "output/Trinity模型报价表.xlsx"

SHEET_COMMERCIAL_MAIN = "10_商务总表-生文"
SHEET_COMMERCIAL_CROSS = "11_交叉模型-生文"
SHEET_IMAGE_MAIN = "20_商务总表-生图"
SHEET_IMAGE_CROSS = "21_交叉模型-生图"
SHEET_VIDEO_MAIN = "30_商务总表-生视频"
SHEET_VIDEO_CROSS = "31_交叉模型-生视频"
SHEET_RESOLUTION = "01_报价解析汇总"

IMAGE_VENDOR_ORDER = [
    "OpenAI",
    "Google",
    "Midjourney",
    "阿里云·通义",
    "腾讯·混元",
    "字节·即梦",
    "可灵",
    "Vidu",
    "其他",
]

VIDEO_VENDOR_ORDER = [
    "Google",
    "字节·即梦",
    "可灵",
    "MiniMax",
    "快手·可灵",
    "Vidu",
    "阿里云·通义",
    "腾讯·混元",
    "其他",
]

# 生图分辨率展示顺序
_IMAGE_SPEC_ORDER = {
    "1K以下": 0,
    "1K": 1,
    "2K": 2,
    "4K": 3,
    "标准": 10,
    "条件": 90,
}

# 本轮写回闸：只上浮这两款目录价（刊例策略 §3.1）；降价项仍用线上
CATALOG_OVERRIDES_USD: dict[str, tuple[float, float, float | None]] = {
    "hy-mt2-plus": (0.1, 0.4, None),
    "glm-5.2": (1.4, 4.4, 0.26),
}

VENDOR_CN = {
    "anthropic": "Anthropic",
    "openai": "OpenAI",
    "google deepMind": "Google",
    "deepseek": "DeepSeek",
    "moonshot": "月之暗面",
    "minimax": "MiniMax",
    "zhipu": "智谱",
    "智谱": "智谱",
    "tongyi": "阿里云·通义",
    "tencent": "腾讯·混元",
    "豆包": "字节·豆包",
    "GPT": "OpenAI",
    "GK": "GK",
}

VENDOR_ORDER = [
    "OpenAI",
    "Anthropic",
    "Google",
    "DeepSeek",
    "月之暗面",
    "MiniMax",
    "智谱",
    "阿里云·通义",
    "腾讯·混元",
    "字节·豆包",
]

# 对外客户档（Standard=目录价列）：仅三档 ≥$5k / ≥$10k / ≥$50k（表头标「对外」）
# 对内商务仍为五档 $1k/$5k/$10k/$30k/$50k；对外折取商务 Mid/Growth/Enterprise
TIERS = [
    ("mid", "对外·≥$5k", 0),
    ("growth", "对外·≥$10k", 1),
    ("enterprise", "对外·≥$50k", 2),
]

# L3a 公开定折：折数如 8.5=八五折；顺序 = $5k / $10k / $50k（浅→深）
# 来源：定价方案-v0 §6；对客折底线 5.5；不直接抄 L3b 全五档
PUBLIC_FAMILY_TIERS: dict[float, list[float]] = {
    # ≤5 折：最深已拍（0.40→5.5；0.50→6.0）；取商务 Mid/Growth/Ent
    0.40: [6.2, 6.0, 5.5],
    0.50: [6.8, 6.5, 6.0],
    0.60: [8.2, 7.6, 6.7],
    0.65: [8.5, 8.2, 7.2],
    0.70: [9.2, 8.8, 7.8],
    0.75: [9.0, 8.7, 8.2],
    0.85: [9.8, 9.7, 9.4],
}

# 单模覆盖（优先于成本族）
PUBLIC_MODEL_TIERS: dict[str, list[float]] = {
    "claude-opus-4-6": [9.5, 9.2, 8.8],
    # 供应侧多为 1.0，对外公开阶梯与 0.65 主锚对齐
    "gpt-5-nano": [8.5, 8.2, 7.2],
}

CLAUDE_OPUS_TIER_ZHE: list[float] = [9.5, 9.2, 8.8]
FAM_078浅折_TIER_ZHE: list[float] = [9.4, 9.2, 8.5]
_LIST_PRICE_ZHE = [10.0, 10.0, 10.0]


def public_tiers_for(model_id: str, fam: float | None) -> list[float | None] | None:
    """Return 3-tier zhe list ($5k/$10k/$50k) for L3a, or None if no public ladder."""
    if model_id in PUBLIC_MODEL_TIERS:
        return list(PUBLIC_MODEL_TIERS[model_id])
    # prefix match for versioned ids
    for mid, zhes in PUBLIC_MODEL_TIERS.items():
        if model_id == mid or model_id.startswith(mid + "-"):
            return list(zhes)
    # 未映射成本族：有刊例外发仍按原价（不用「—」）
    if fam is None:
        return list(_LIST_PRICE_ZHE)
    # 口径：Claude 只允许在成本族=1.0 时显示为原价；
    # 其它成本族的 Claude 全部使用与 opus-4-6 相同的公开折扣梯度。
    if str(model_id).startswith("claude-"):
        if abs(float(fam) - 1.0) < 1e-9:
            return list(_LIST_PRICE_ZHE)
        return list(CLAUDE_OPUS_TIER_ZHE)
    if fam in PUBLIC_FAMILY_TIERS:
        return list(PUBLIC_FAMILY_TIERS[fam])
    # 0.78：如果模型没有更低折扣族可选，则给一个“浅折”公开梯度
    if abs(float(fam) - 0.78) < 1e-6:
        return list(FAM_078浅折_TIER_ZHE)
    # 1.0 及其它 ≥0.78 未进公开分族：对外标原价
    if fam >= 0.78:
        return list(_LIST_PRICE_ZHE)
    return None

def parse_usd(s):
    """Parse 入$x · 出$y · 缓$z → (in, out, cache|None)."""
    if not s or s == "—":
        return None
    m = re.search(
        r"入\s*\$([\d.]+)\s*·\s*出\s*\$([\d.]+)(?:\s*·\s*缓\s*\$([\d.]+))?",
        str(s),
    )
    if not m:
        return None
    cache = float(m.group(3)) if m.group(3) else None
    return float(m.group(1)), float(m.group(2)), cache


def ceil_money_2(v: float | None) -> float | None:
    """对外金额：最多两位小数；第 3 位起有数则向上进到分。

    例：0.615→0.62，0.610→0.61，1.2→1.2，2→2
    """
    if v is None:
        return None
    x = float(v)
    if not math.isfinite(x):
        return None
    # 浮点误差：略小于整分时不误进
    scaled = x * 100.0
    if abs(scaled - round(scaled)) < 1e-9:
        return round(scaled) / 100.0
    return math.ceil(scaled - 1e-12) / 100.0


def fmt_num(v: float) -> str:
    """对外展示固定两位小数（已 ceil 进位）。"""
    x = ceil_money_2(v)
    if x is None:
        return "—"
    return f"{x:.2f}"


def fmt_cache(v: float | None) -> str:
    if v is None:
        return "—"
    return fmt_num(v)


def fmt_tier_zhe(zhe: float | None) -> str:
    """阶梯列只展示对客折：如 8.5折 / 原价（入出同折，不写绝对价）。"""
    if zhe is None:
        return "—"
    if zhe >= 9.999:
        return "原价"
    zhe_s = (
        f"{zhe:.1f}".rstrip("0").rstrip(".")
        if abs(zhe - round(zhe)) > 1e-9
        else str(int(round(zhe)))
    )
    return f"{zhe_s}折"


def parse_zhe_cell(val) -> float | None:
    """Parse '9.0（GM 28%）' / '原价（GM 25%）' / '—' → 折数 or 10.0 or None."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s == "—" or s == "-":
        return None
    if s.startswith("原价"):
        return 10.0
    m = re.match(r"([\d.]+)", s)
    if not m:
        return None
    return float(m.group(1))


def parse_family_key(label: str) -> float | None:
    m = re.match(r"(0\.\d+|1\.0)", str(label))
    return float(m.group(1)) if m else None


def _ingest_commercial_pair(
    wb,
    main_sheet: str,
    cross_sheet: str,
    family_tiers: dict[float, list[float | None]],
    model_families: dict[str, list[float]],
    recommended: dict[str, float],
    required: bool = True,
) -> int:
    """Parse one 商务总表 + 交叉模型 pair into shared maps. Returns model count added."""
    if main_sheet not in wb.sheetnames:
        if required:
            raise ValueError(
                f"{SRC_COMMERCIAL.name} 缺少 {main_sheet}；"
                "请先跑 rebuild_discount_tier_workbook.py"
            )
        return 0
    # 占位页（待补）跳过
    a1 = str(wb[main_sheet]["A1"].value or "")
    if "待补" in a1:
        return 0

    before = len(model_families)
    ws = wb[main_sheet]
    for row in ws.iter_rows(min_row=2, max_row=20, max_col=11, values_only=True):
        if not row[0]:
            continue
        fam = parse_family_key(row[0])
        if fam is None:
            continue
        tiers = [parse_zhe_cell(row[i]) for i in range(1, 6)]
        family_tiers.setdefault(fam, tiers)
        models_cell = row[8]
        if not models_cell or str(models_cell).strip() in ("（待补）", "—", "-"):
            continue
        for part in re.split(r"[；;\n]+", str(models_cell)):
            part = part.strip()
            if not part:
                continue
            mid = part.split("：")[0].split(":")[0].strip().rstrip("※").strip()
            if not mid:
                continue
            model_families.setdefault(mid, [])
            if fam not in model_families[mid]:
                model_families[mid].append(fam)

    if cross_sheet in wb.sheetnames and "待补" not in str(wb[cross_sheet]["A1"].value or ""):
        ws2 = wb[cross_sheet]
        header_row = None
        rec_col = None
        for r in range(1, 6):
            for c in range(1, ws2.max_column + 1):
                if str(ws2.cell(r, c).value or "").strip() == "推荐成本折":
                    header_row = r
                    rec_col = c
                    break
            if rec_col:
                break
        if rec_col is not None:
            for r in range(header_row + 1, ws2.max_row + 1):
                mid = ws2.cell(r, 1).value
                rec = ws2.cell(r, rec_col).value
                if not mid or rec in (None, ""):
                    continue
                try:
                    recommended[str(mid).strip()] = float(rec)
                except (TypeError, ValueError):
                    continue
    return len(model_families) - before


def load_commercial():
    """Return family_tiers, model_family[id], recommended（10/11+20/21+30/31）。"""
    if not SRC_COMMERCIAL.exists():
        raise FileNotFoundError(f"商务表缺失: {SRC_COMMERCIAL}")
    wb = openpyxl.load_workbook(SRC_COMMERCIAL, data_only=True)
    family_tiers: dict[float, list[float | None]] = {}
    model_families: dict[str, list[float]] = {}
    recommended: dict[str, float] = {}

    n_text = _ingest_commercial_pair(
        wb,
        SHEET_COMMERCIAL_MAIN,
        SHEET_COMMERCIAL_CROSS,
        family_tiers,
        model_families,
        recommended,
        required=True,
    )
    n_img = _ingest_commercial_pair(
        wb,
        SHEET_IMAGE_MAIN,
        SHEET_IMAGE_CROSS,
        family_tiers,
        model_families,
        recommended,
        required=False,
    )
    n_vid = _ingest_commercial_pair(
        wb,
        SHEET_VIDEO_MAIN,
        SHEET_VIDEO_CROSS,
        family_tiers,
        model_families,
        recommended,
        required=False,
    )

    print(
        f"commercial {SRC_COMMERCIAL.name}: "
        f"models={len(model_families)} (+text≈{n_text} +image≈{n_img} +video≈{n_vid}) "
        f"families={sorted(family_tiers)} recommended={len(recommended)}"
    )
    return family_tiers, model_families, recommended


def resolve_family(
    model_id: str,
    model_families: dict[str, list[float]],
    recommended: dict[str, float],
) -> float | None:
    if model_id in recommended:
        return recommended[model_id]
    fams = model_families.get(model_id)
    if not fams:
        # try without -old / -preview suffixes already exact
        return None
    # 厚利优先：成本折数字越小越优先
    return min(fams)


def _num_amount(block) -> float | None:
    if not isinstance(block, dict):
        return None
    try:
        return float(block.get("amount"))
    except (TypeError, ValueError):
        return None


def _online_default_prices(entry: dict) -> tuple[float, float, float | None] | None:
    """Take first usable default / short-context tier from /v1/prices entry.

    Supports legacy ``default`` groups and ``token_kind`` + ranges
    (context_length_range / usage_token_range).
    """
    groups = entry.get("price_groups") or []

    # 1) legacy / flat: prices.input + prices.output on a group
    ranked = []
    for g in groups:
        prices = g.get("prices") or {}
        inp = _num_amount(prices.get("input"))
        out = _num_amount(prices.get("output"))
        if inp is None or out is None:
            continue
        cache = _num_amount(prices.get("cache")) or _num_amount(
            prices.get("cached_input")
        )
        label = str(g.get("label") or g.get("conditions_summary") or "")
        demote = 1 if re.search(r">\s*272|>\s*262|长上下文|long", label, re.I) else 0
        ranked.append((demote, g.get("type") != "default", inp, out, cache))
    if ranked:
        ranked.sort(key=lambda x: (x[0], x[1]))
        _, _, inp, out, cache = ranked[0]
        return inp, out, cache

    # 2) token_kind ranges (GPT-5.x / Gemini pro / GLM / Qwen 分档)
    def _kind(kind: str):
        return next(
            (
                g
                for g in groups
                if g.get("type") == "token_kind" and g.get("token_kind") == kind
            ),
            None,
        )

    input_g = _kind("input")
    out_g = _kind("output")
    cache_g = _kind("cached_input") or _kind("cache")
    ranges = (input_g or {}).get("ranges") or []
    if not ranges:
        return None

    def range_score(r: dict) -> tuple[int, float]:
        label = str(
            (r.get("range") or {}).get("display_short")
            or (r.get("range") or {}).get("display")
            or ""
        )
        demote = 1 if re.search(r"^>|长|long", label, re.I) else 0
        mn = (r.get("range") or {}).get("min")
        try:
            min_v = float(mn) if mn is not None else 0.0
        except (TypeError, ValueError):
            min_v = 0.0
        return demote, min_v

    order = sorted(range(len(ranges)), key=lambda i: range_score(ranges[i]))
    i = order[0]
    inp = _num_amount((ranges[i].get("price") or {}))
    out_ranges = (out_g or {}).get("ranges") or []
    cache_ranges = (cache_g or {}).get("ranges") or []
    out = _num_amount((out_ranges[i].get("price") or {})) if i < len(out_ranges) else None
    cache = (
        _num_amount((cache_ranges[i].get("price") or {}))
        if i < len(cache_ranges)
        else None
    )
    if inp is None or out is None:
        return None
    return inp, out, cache


def _guess_vendor(model_id: str, display_name: str = "") -> str:
    mid = (model_id or "").lower()
    name = (display_name or "").lower()
    if mid.startswith("gpt-") or mid.startswith("o1") or mid.startswith("o3"):
        return "OpenAI"
    if mid.startswith("claude") or "anthropic" in name:
        return "Anthropic"
    if mid.startswith("gemini") or "google" in name:
        return "Google"
    if mid.startswith("deepseek"):
        return "DeepSeek"
    if mid.startswith("kimi") or "moonshot" in name:
        return "月之暗面"
    if mid.startswith("minimax"):
        return "MiniMax"
    if mid.startswith("glm") or "智谱" in name or "zhipu" in name:
        return "智谱"
    if mid.startswith("qwen"):
        return "阿里云·通义"
    if mid.startswith("hy") or mid.startswith("hunyuan"):
        return "腾讯·混元"
    if mid.startswith("doubao") or "豆包" in name:
        return "字节·豆包"
    if mid.startswith("grok"):
        return "GK"
    return "其他"


def _guess_vendor_image(model_id: str, display_name: str = "") -> str:
    mid = (model_id or "").lower()
    name = (display_name or "").lower()
    if mid.startswith("gpt-") or mid.startswith("og-image") or "image-2" in name:
        return "OpenAI"
    if mid.startswith("gemini") or mid.startswith("gg-") or "nano" in name:
        return "Google"
    if mid.startswith("mj") or "midjourney" in name:
        return "Midjourney"
    if mid.startswith("qwen"):
        return "阿里云·通义"
    if mid.startswith("hunyuan") or mid.startswith("hy"):
        return "腾讯·混元"
    if mid.startswith("jimeng") or mid.startswith("si-") or "seedream" in name:
        return "字节·即梦"
    if mid.startswith("kling"):
        return "可灵"
    if mid.startswith("vidu"):
        return "Vidu"
    return "其他"


def _norm_image_spec(label: str) -> str:
    s = str(label or "").strip()
    sl = s.lower().replace(" ", "")
    if sl in ("lt_1k", "lt1k", "<1k", "sub1k") or "1k以下" in s:
        return "1K以下"
    if re.fullmatch(r"1k", sl) or s in ("1K", "1k"):
        return "1K"
    if re.fullmatch(r"2k", sl) or s in ("2K", "2k"):
        return "2K"
    if re.fullmatch(r"4k", sl) or s in ("4K", "4k"):
        return "4K"
    if "条件" in s:
        return "条件"
    if not s or s in ("标准计价", "标准价", "default"):
        return "标准"
    return s


def load_image_rows() -> list[dict]:
    """生图：线上 /v1/prices image；按张分分辨率行，按 token 则一行汇总目录价。"""
    if not SRC_ONLINE_IMAGE.exists():
        raise FileNotFoundError(f"missing online image prices: {SRC_ONLINE_IMAGE}")
    raw = json.loads(SRC_ONLINE_IMAGE.read_text(encoding="utf-8"))
    entries = raw.get("data") or raw.get("models") or []
    rows: list[dict] = []
    for e in entries:
        mid = str(e.get("model") or e.get("id") or "").strip()
        if not mid:
            continue
        display = str(e.get("display_name") or mid).strip()
        vendor = _guess_vendor_image(mid, display)
        charge = str(e.get("charge_unit") or "")
        groups = e.get("price_groups") or []

        if charge == "image_count" or str(e.get("price_unit") or "") == "per_image":
            seen: set[tuple[str, float]] = set()
            for g in groups:
                prices = g.get("prices") or {}
                unit_block = prices.get("unit") or prices.get("image") or {}
                amt = _num_amount(unit_block)
                if amt is None:
                    continue
                spec = _norm_image_spec(
                    g.get("label") or g.get("type") or g.get("conditions_summary") or ""
                )
                key = (spec, round(amt, 6))
                if key in seen:
                    continue
                seen.add(key)
                rows.append(
                    {
                        "vendor": vendor,
                        "model_id": mid,
                        "display": display,
                        "spec": spec,
                        "unit": "USD/张",
                        "catalog": ceil_money_2(amt),
                        "catalog_display": None,
                    }
                )
        else:
            # token 生图：一行列出入/出/缓
            prices = _online_default_prices(e)
            if not prices:
                continue
            inp, outp, cache = prices
            parts = [f"入 {fmt_num(inp)}", f"出 {fmt_num(outp)}"]
            if cache is not None:
                parts.append(f"缓 {fmt_num(cache)}")
            rows.append(
                {
                    "vendor": vendor,
                    "model_id": mid,
                    "display": display,
                    "spec": "标准",
                    "unit": "USD/百万 tokens",
                    "catalog": ceil_money_2(inp),
                    "catalog_display": " · ".join(parts),
                }
            )

    order = {v: i for i, v in enumerate(IMAGE_VENDOR_ORDER)}

    def sort_key(r: dict):
        return (
            order.get(r["vendor"], 99),
            r["model_id"],
            _IMAGE_SPEC_ORDER.get(r["spec"], 50),
            r["spec"],
        )

    rows.sort(key=sort_key)
    return rows


def has_public_ladder(
    model_id: str,
    model_families: dict[str, list[float]],
    recommended: dict[str, float],
) -> bool:
    fam = resolve_family(model_id, model_families, recommended)
    zhes = public_tiers_for(model_id, fam)
    return bool(zhes and any(z is not None and z < 9.999 for z in zhes))


def _dedupe_models_by_id(rows: list[dict]) -> list[dict]:
    out: list[dict] = []
    seen: set[str] = set()
    for r in rows:
        mid = r["model_id"]
        if mid in seen:
            continue
        seen.add(mid)
        out.append(r)
    return out


def _listing_cols_for_overview(mod: str, m: dict) -> tuple[str, str, str]:
    """对外折扣一览刊例三列（与 01 生文表头对齐）；图/视频多规格时输入列给摘要。"""
    if mod == "生文":
        return (
            fmt_num(m.get("inp")),
            fmt_num(m.get("out")),
            fmt_cache(m.get("cache")),
        )
    # 生图/生视频：token 汇总或代表价放「输入」；出/缓无则 —
    disp = m.get("catalog_display")
    if disp:
        return (str(disp), "—", "—")
    cat = m.get("catalog")
    unit = str(m.get("unit") or "")
    spec = str(m.get("spec") or "")
    if cat is not None:
        label = fmt_num(cat)
        if unit or spec:
            label = f"{label}" + (f"（{spec}）" if spec else "")
        return (label, "—", "—")
    return ("—", "—", "—")


def collect_discounted_models(
    models: list[dict],
    image_rows: list[dict],
    video_rows: list[dict],
    model_families: dict[str, list[float]],
    recommended: dict[str, float],
) -> list[dict]:
    """有公开用量折扣的模型一览（原价不进）；按模态→最深折→厂商。不含成本族。"""
    modality_items = (
        ("生文", models),
        ("生图", _dedupe_models_by_id(image_rows)),
        ("生视频", _dedupe_models_by_id(video_rows)),
    )
    rows: list[dict] = []
    for mod, items in modality_items:
        for m in items:
            mid = m["model_id"]
            if not has_public_ladder(mid, model_families, recommended):
                continue
            fam = resolve_family(mid, model_families, recommended)
            zhes = public_tiers_for(mid, fam) or [None, None, None]
            deep = min((z for z in zhes if z is not None), default=10.0)
            inp, outp, cache = _listing_cols_for_overview(mod, m)
            rows.append(
                {
                    "modality": mod,
                    "vendor": m.get("vendor") or "",
                    "model_id": mid,
                    "display": m.get("display") or mid,
                    "inp": inp,
                    "out": outp,
                    "cache": cache,
                    "zhes": zhes,
                    "deep": deep,
                }
            )
    mod_order = {"生文": 0, "生图": 1, "生视频": 2}
    rows.sort(
        key=lambda r: (
            mod_order.get(r["modality"], 9),
            r["deep"],
            r["vendor"],
            r["model_id"],
        )
    )
    return rows


def write_discount_overview_sheet(
    wb,
    discount_rows: list[dict],
    *,
    valid_until: str,
    thin,
    white_bold,
    body_font,
    num_font,
    center,
    left,
) -> None:
    """00_折扣一览：仅有折扣模型；表头对齐 01（+模态）；无成本信息。"""
    name = "00_折扣一览"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 0)
    fills = {
        "生文": solid_fill("DBEAFE"),  # 蓝
        "生图": solid_fill("FEF3C7"),  # 琥珀
        "生视频": solid_fill("D1FAE5"),  # 绿
    }
    header_fill = solid_fill("1B4F72")
    n_cols = 7 + len(TIERS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws["A1"] = "Trinity · 对外折扣一览（仅列有用量折扣的模型）"
    ws["A1"].font = Font(name="PingFang SC", size=16, bold=True, color="1D2939")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    counts = {
        m: sum(1 for r in discount_rows if r["modality"] == m)
        for m in ("生文", "生图", "生视频")
    }
    ws["A2"] = (
        f"有效期：{valid_until}　|　共 {len(discount_rows)} 款有折扣"
        f"（生文 {counts['生文']} · 生图 {counts['生图']} · 生视频 {counts['生视频']}）。"
        "底色：生文蓝 · 生图琥珀 · 生视频绿。"
        "表头与分册一致（+模态）；不含成本/线路。"
        "图/视频多规格刊例见 02/03；达档=企业户累积消耗（按刊例价计）。"
    )
    ws["A2"].font = Font(name="PingFang SC", size=10, color="475467")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 40

    headers = [
        "模态",
        "厂商",
        "模型 ID",
        "显示名",
        "刊例价·输入",
        "刊例价·输出",
        "刊例价·缓存",
    ] + [label for _, label, _ in TIERS]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(4, i, h)
        style_cell(cell, fill=header_fill, font=white_bold, align=center, border=thin)
    ws.row_dimensions[4].height = 24

    for idx, row in enumerate(discount_rows):
        r = 5 + idx
        fill = fills.get(row["modality"])
        zhes = row["zhes"]
        vals = [
            row["modality"],
            row["vendor"],
            row["model_id"],
            row["display"],
            row["inp"],
            row["out"],
            row["cache"],
        ]
        for ti in range(len(TIERS)):
            vals.append(fmt_tier_zhe(zhes[ti] if ti < len(zhes) else None))
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            style_cell(
                cell,
                font=num_font if c >= 5 else body_font,
                align=center if c == 1 or c >= 5 else left,
                border=thin,
                fill=fill,
            )

    for i, w in enumerate([8, 12, 26, 22, 14, 14, 12] + [14] * len(TIERS), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "H5"
    print(
        f"discount overview: {len(discount_rows)} models "
        f"(text={counts['生文']} image={counts['生图']} video={counts['生视频']})"
    )


def sort_image_rows_discount_first(
    rows: list[dict],
    model_families: dict[str, list[float]],
    recommended: dict[str, float],
) -> list[dict]:
    """有公开用量折扣的模型整组置顶，同模规格仍按分辨率序。"""
    order = {v: i for i, v in enumerate(IMAGE_VENDOR_ORDER)}
    ladder_cache: dict[str, bool] = {}

    def has_ladder(mid: str) -> bool:
        if mid not in ladder_cache:
            ladder_cache[mid] = has_public_ladder(mid, model_families, recommended)
        return ladder_cache[mid]

    def sort_key(r: dict):
        return (
            0 if has_ladder(r["model_id"]) else 1,
            order.get(r["vendor"], 99),
            r["model_id"],
            _IMAGE_SPEC_ORDER.get(r["spec"], 50),
            r["spec"],
        )

    return sorted(rows, key=sort_key)


def _guess_vendor_video(model_id: str, display_name: str = "") -> str:
    mid = (model_id or "").lower()
    name = (display_name or "").lower()
    if mid.startswith("gemini") or mid.startswith("gv-"):
        return "Google"
    if mid.startswith("seedance") or mid.startswith("jimeng") or mid.startswith("jv-"):
        return "字节·即梦"
    if mid.startswith("kling"):
        return "可灵"
    if mid.startswith("hailuo") or mid.startswith("minimax") or mid.startswith("h2"):
        return "MiniMax"
    if mid.startswith("happyhorse"):
        return "其他"
    if mid.startswith("vidu"):
        return "Vidu"
    if mid.startswith("wan") or mid.startswith("os-"):
        return "阿里云·通义"
    if mid.startswith("pixverse"):
        return "其他"
    if "hunyuan" in mid or mid.startswith("hy"):
        return "腾讯·混元"
    return "其他"


def _norm_video_spec(label: str) -> str:
    s = str(label or "").strip()
    if not s:
        return "标准"
    sl = s.lower().replace(" ", "")
    for key in ("4k", "1080p", "720p", "768p", "480p", "540p"):
        if key in sl:
            return key.upper() if key == "4k" else key
    if "输入" in s or "input" in sl:
        return "输入"
    if "视频输出" in s or "output_video" in sl or "video" in sl and "输出" in s:
        return "视频输出"
    if "文本" in s or "output_text" in sl or "思考" in s:
        return "文本/思考输出"
    return s[:24]


_VIDEO_SPEC_ORDER = {
    "输入": 0,
    "文本/思考输出": 1,
    "视频输出": 2,
    "480p": 10,
    "540p": 11,
    "720p": 12,
    "768p": 13,
    "1080p": 14,
    "4K": 15,
    "标准": 20,
}


def load_video_rows() -> list[dict]:
    """生视频：线上刊例；按分辨率/token 种类分行。"""
    if not SRC_ONLINE_VIDEO.exists():
        raise FileNotFoundError(f"missing online video prices: {SRC_ONLINE_VIDEO}")
    raw = json.loads(SRC_ONLINE_VIDEO.read_text(encoding="utf-8"))
    entries = raw.get("data") or raw.get("models") or []
    rows: list[dict] = []
    for e in entries:
        mid = str(e.get("model") or e.get("id") or "").strip()
        if not mid:
            continue
        display = str(e.get("display_name") or mid).strip()
        vendor = _guess_vendor_video(mid, display)
        charge = str(e.get("charge_unit") or "")
        groups = e.get("price_groups") or []
        seen: set[tuple[str, str, float]] = set()

        for g in groups:
            prices = g.get("prices") or {}
            # unit / per-second / per-resolution
            unit_block = prices.get("unit") or prices.get("video") or {}
            amt = _num_amount(unit_block)
            kind = None
            if amt is None:
                for k in ("input", "output_text", "output_video", "output"):
                    if k in prices:
                        amt = _num_amount(prices.get(k))
                        kind = k
                        break
            if amt is None:
                continue
            label = g.get("label") or g.get("type") or g.get("conditions_summary") or ""
            if kind == "input":
                spec = "输入"
            elif kind == "output_text":
                spec = "文本/思考输出"
            elif kind == "output_video":
                spec = "视频输出"
            else:
                spec = _norm_video_spec(str(label))
            key = (spec, charge, round(float(amt), 6))
            if key in seen:
                continue
            seen.add(key)
            if charge == "video_token" and kind:
                unit = "USD/百万 video tokens"
            elif charge == "video_token":
                unit = "USD/百万 video tokens"
            else:
                unit = "USD/秒"
            rows.append(
                {
                    "vendor": vendor,
                    "model_id": mid,
                    "display": display,
                    "spec": spec,
                    "unit": unit,
                    "catalog": ceil_money_2(float(amt)),
                    "catalog_display": None,
                }
            )

    order = {v: i for i, v in enumerate(VIDEO_VENDOR_ORDER)}

    def sort_key(r: dict):
        return (
            order.get(r["vendor"], 99),
            r["model_id"],
            _VIDEO_SPEC_ORDER.get(r["spec"], 50),
            r["spec"],
        )

    rows.sort(key=sort_key)
    return rows


def sort_video_rows_discount_first(
    rows: list[dict],
    model_families: dict[str, list[float]],
    recommended: dict[str, float],
) -> list[dict]:
    order = {v: i for i, v in enumerate(VIDEO_VENDOR_ORDER)}
    ladder_cache: dict[str, bool] = {}

    def has_ladder(mid: str) -> bool:
        if mid not in ladder_cache:
            ladder_cache[mid] = has_public_ladder(mid, model_families, recommended)
        return ladder_cache[mid]

    def sort_key(r: dict):
        return (
            0 if has_ladder(r["model_id"]) else 1,
            order.get(r["vendor"], 99),
            r["model_id"],
            _VIDEO_SPEC_ORDER.get(r["spec"], 50),
            r["spec"],
        )

    return sorted(rows, key=sort_key)


def load_compare_meta() -> dict[str, dict]:
    """Trinity ID → vendor/display from 刊例对比（可选 enrichment）。"""
    if not SRC_PRICES.exists():
        return {}
    wb = openpyxl.load_workbook(SRC_PRICES, data_only=True)
    if "刊例对比校验-生文" not in wb.sheetnames:
        return {}
    ws = wb["刊例对比校验-生文"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(h or "").strip() for h in rows[0]]
    try:
        i_tid = header.index("Trinity ID")
        i_name = header.index("显示名")
        i_brand = header.index("厂商")
    except ValueError:
        return {}
    meta: dict[str, dict] = {}
    for r in rows[1:]:
        if not r or not r[i_tid] or r[i_tid] in ("—", "-"):
            continue
        tid = str(r[i_tid]).strip()
        if tid in meta:
            continue
        brand = str(r[i_brand] or "").strip()
        meta[tid] = {
            "display": str(r[i_name] or tid).strip(),
            "vendor": VENDOR_CN.get(brand, brand or "其他"),
        }
    return meta


def load_models():
    """目录价主源 = 线上 /v1/prices（全量上架）；一张模型一行（取短上下文/默认档）。"""
    if not SRC_ONLINE.exists():
        raise FileNotFoundError(f"missing online prices: {SRC_ONLINE}")
    raw = json.loads(SRC_ONLINE.read_text(encoding="utf-8"))
    entries = raw.get("data") or raw.get("models") or []
    meta = load_compare_meta()
    out = []
    for e in entries:
        mid = str(e.get("model") or e.get("id") or "").strip()
        if not mid:
            continue
        prices = _online_default_prices(e)
        if not prices:
            continue
        inp, outp, cache = prices
        if mid in CATALOG_OVERRIDES_USD:
            oi, oo, oc = CATALOG_OVERRIDES_USD[mid]
            inp, outp = oi, oo
            if oc is not None:
                cache = oc
        m = meta.get(mid, {})
        display = m.get("display") or e.get("display_name") or mid
        vendor = m.get("vendor") or _guess_vendor(mid, display)
        out.append(
            {
                "vendor": vendor,
                "model_id": mid,
                "display": display,
                "tier": "",
                "inp": ceil_money_2(inp),
                "out": ceil_money_2(outp),
                "cache": ceil_money_2(cache),
            }
        )
    order = {v: i for i, v in enumerate(VENDOR_ORDER)}
    out.sort(key=lambda m: (order.get(m["vendor"], 99), m["model_id"]))
    return out


def style_cell(cell, *, fill=None, font=None, align=None, border=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border


def _sheet_styles():
    thin = Border(
        left=Side(style="thin", color="E4E7EC"),
        right=Side(style="thin", color="E4E7EC"),
        top=Side(style="thin", color="E4E7EC"),
        bottom=Side(style="thin", color="E4E7EC"),
    )
    return {
        "thin": thin,
        "header_fill": solid_fill("1B4F72"),
        "alt": solid_fill("F8FAFC"),
        "body_font": Font(name="PingFang SC", size=10),
        "num_font": Font(name="Menlo", size=9),
        "white_bold": Font(name="PingFang SC", size=10, bold=True, color="FFFFFF"),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center"),
    }


def ladder_source_label(
    model_id: str,
    fam: float | None,
    recommended: dict[str, float],
) -> str:
    if str(model_id).startswith("claude-") and (
        fam is None or abs(float(fam) - 1.0) > 1e-9
    ):
        src = "Claude对齐opus阶梯"
    elif model_id == "gpt-5-nano" or model_id.startswith("gpt-5-nano-"):
        src = "单模对齐0.65公开阶梯"
    elif model_id in PUBLIC_MODEL_TIERS or any(
        model_id == mid or model_id.startswith(mid + "-")
        for mid in PUBLIC_MODEL_TIERS
    ):
        src = "单模特例"
    elif fam in PUBLIC_FAMILY_TIERS:
        src = f"公开族 {fam}"
    elif fam is not None and abs(float(fam) - 0.78) < 1e-6:
        src = "0.78浅折"
    elif fam is not None and abs(float(fam) - 1.0) < 1e-9:
        src = "1.0→原价"
    elif fam is not None and fam >= 0.78:
        src = "≥0.78→原价"
    elif fam is not None:
        src = "族未进公开矩阵"
    else:
        src = "未映射→原价"
    if model_id in recommended:
        return f"{src}（02推荐）"
    if fam is not None and not src.startswith("未映射"):
        return f"{src}（01归属）"
    return src


def _parse_cost_ratios(cost_cell) -> list[float]:
    ratios: list[float] = []
    if not cost_cell:
        return ratios
    for m in re.finditer(r"官方\s*([\d.]+)\s*/\s*成本\s*([\d.]+)", str(cost_cell)):
        o, c = float(m.group(1)), float(m.group(2))
        if o > 0:
            ratios.append(c / o)
    return ratios


def scan_route_lines() -> list[dict]:
    """扫描生文+生图+生视频线路源：启用+停用，含成本/官方比。"""
    rows: list[dict] = []
    for root, tag in (
        (ROUTES_TEXT, "text"),
        (ROUTES_IMAGE, "image"),
        (ROUTES_VIDEO, "video"),
    ):
        if not root.exists():
            continue
        for path in sorted(root.glob("*.xlsx")):
            if path.name.startswith("~") or path.name.lower() == "readme.md":
                continue
            wb = openpyxl.load_workbook(path, data_only=True)
            if "线路管理" not in wb.sheetnames:
                continue
            ws = wb["线路管理"]
            cur = None
            for r in range(2, ws.max_row + 1):
                mid = ws.cell(r, 1).value
                if mid:
                    cur = str(mid).strip()
                if not cur:
                    continue
                route = ws.cell(r, 4).value
                if not mid and not route and not ws.cell(r, 9).value:
                    continue
                disc = ws.cell(r, 9).value
                en = ws.cell(r, 11).value
                cost = ws.cell(r, 12).value
                ratios = _parse_cost_ratios(cost)
                if not mid and not route:
                    continue
                rows.append(
                    {
                        "source": f"{tag}/{path.name}",
                        "modality": tag,
                        "model_id": cur,
                        "route": str(route or "").strip(),
                        "discount": disc,
                        "status": str(en or "").strip(),
                        "min_ratio": min(ratios) if ratios else None,
                        "ratios": ratios,
                    }
                )
    return rows


def write_price_resolution_sheet(
    wb,
    meta: list[tuple[str, str]],
    models: list[dict],
    image_rows: list[dict],
    video_rows: list[dict],
    model_families: dict[str, list[float]],
    recommended: dict[str, float],
) -> None:
    """01_报价解析汇总：对外阶梯依据（含停用更低进价；吸收原 90/91）。"""
    name = SHEET_RESOLUTION
    # 保持页签顺序：删后插到 index 1（00 之后）
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 1)
    route_rows = scan_route_lines()
    by_mid: dict[str, list[dict]] = {}
    for rr in route_rows:
        by_mid.setdefault(rr["model_id"], []).append(rr)

    head_font = Font(name="PingFang SC", size=14, bold=True, color="B42318")
    sec_font = Font(name="PingFang SC", size=11, bold=True, color="1D2939")
    bold = Font(name="PingFang SC", bold=True)
    body = Font(name="PingFang SC", size=10)
    mono = Font(name="Menlo", size=9)
    hint = Font(name="PingFang SC", size=9, color="64748B", italic=True)
    fill_head = solid_fill("FEF3C7")
    fill_sec = solid_fill("F1F5F9")
    fill_yuan = solid_fill("FEF2F2")
    fill_ok = solid_fill("ECFDF3")

    ws["A1"] = "01_报价解析汇总 · 对外阶梯依据（对内 · 勿外发）"
    ws["A1"].font = head_font
    ws.merge_cells("A1:L1")
    ws["A2"] = (
        f"对应外发：{OUT.name}（01_生文/02_生图/03_生视频）｜"
        f"线路源：{ROUTES_TEXT.name}/ · {ROUTES_IMAGE.name}/ · {ROUTES_VIDEO.name}/｜"
        "启用线路→商务成本族→公开阶梯；停用更低进价仅作依据，不自动改对外折。"
        "重建：python3 pricing/scripts/build_outward_quote_standard.py"
    )
    ws["A2"].font = hint
    ws.merge_cells("A2:L2")
    ws.row_dimensions[2].height = 36

    # 外发口径摘要（原 90）
    ws["A3"] = "〇、外发口径摘要"
    ws["A3"].font = sec_font
    ws["A3"].fill = fill_sec
    ws.merge_cells("A3:L3")
    r_meta = 4
    for k, v in meta[:10]:
        ws.cell(r_meta, 1, k).font = bold
        ws.cell(r_meta, 2, v).font = body
        ws.cell(r_meta, 2).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r_meta, start_column=2, end_row=r_meta, end_column=6)
        r_meta += 1
    r_concl = r_meta + 1

    # ---- 结论条 ----
    yuanjia_models = []
    for m in models:
        fam = resolve_family(m["model_id"], model_families, recommended)
        zhes = public_tiers_for(m["model_id"], fam)
        is_yuan = bool(zhes) and all(z is not None and z >= 9.999 for z in zhes)
        if is_yuan:
            yuanjia_models.append(m["model_id"])

    dormant_cheaper = []
    for mid in yuanjia_models:
        for rr in by_mid.get(mid, []):
            st = rr["status"]
            if st == "启用":
                continue
            ratio = rr["min_ratio"]
            if ratio is not None and ratio < 0.97:
                dormant_cheaper.append((mid, rr))

    ws.cell(r_concl, 1, "一、结论（原价对外）").font = sec_font
    ws.cell(r_concl, 1).fill = fill_sec
    ws.merge_cells(start_row=r_concl, start_column=1, end_row=r_concl, end_column=12)
    conclusions = [
        (
            "对外原价款数",
            f"{len(yuanjia_models)} / {len(models)}（生文）",
        ),
        (
            "原价原因",
            "当前启用线路折扣列=1 且成本/官方≈1.0；商务归属 1.0 族 → 公开不设用量阶梯",
        ),
        (
            "是否「有折扣进价却对外原价」",
            "否（启用侧无厚利）；有停用更低进价见第三节",
        ),
        (
            "停用更低进价条数",
            f"{len(dormant_cheaper)} 条（涉及 "
            f"{len({m for m, _ in dormant_cheaper})} 款原价模型）",
        ),
    ]
    for i, (k, v) in enumerate(conclusions):
        r = r_concl + 1 + i
        ws.cell(r, 1, k).font = bold
        ws.cell(r, 2, v).font = body
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

    # ---- 第二节：全量生文解析 ----
    r0 = r_concl + 1 + len(conclusions) + 1
    ws.cell(r0, 1, "二、生文全量 · 报价解析（目录价阶梯依据）").font = sec_font
    ws.cell(r0, 1).fill = fill_sec
    ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=12)
    headers = [
        "厂商",
        "模型 ID",
        "显示名",
        "映射成本折",
        "对外·≥$5k",
        "对外·≥$10k",
        "对外·≥$50k",
        "阶梯来源",
        "启用最低成本比",
        "启用线路摘要",
        "停用更低进价",
        "判定",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r0 + 1, c, h)
        cell.font = bold
        cell.fill = fill_head

    def emit_resolution_rows(start_r: int, items: list[dict]) -> int:
        rr = start_r
        for m in items:
            mid = m["model_id"]
            fam = resolve_family(mid, model_families, recommended)
            zhes = public_tiers_for(mid, fam) or [None, None, None]
            src = ladder_source_label(mid, fam, recommended)
            enabled = [x for x in by_mid.get(mid, []) if x["status"] == "启用"]
            disabled_cheap = [
                x
                for x in by_mid.get(mid, [])
                if x["status"] != "启用"
                and x["min_ratio"] is not None
                and x["min_ratio"] < 0.97
            ]
            en_ratios = [x["min_ratio"] for x in enabled if x["min_ratio"] is not None]
            min_en = min(en_ratios) if en_ratios else None
            en_routes = "；".join(
                dict.fromkeys(
                    f"{x['route'] or '—'}"
                    + (f"({x['min_ratio']:.2f})" if x["min_ratio"] is not None else "")
                    for x in enabled
                    if x["route"]
                )
            )[:120]
            dis_txt = "；".join(
                f"{x['route'] or '—'}·{x['discount']}·{x['status']}"
                f"·比{x['min_ratio']:.2f}·{x['source']}"
                for x in disabled_cheap
            )[:160]
            is_yuan = all(z is not None and z >= 9.999 for z in zhes)
            if is_yuan and min_en is not None and min_en >= 0.97:
                judge = "启用进价≈原价 → 对外原价"
            elif is_yuan and disabled_cheap:
                judge = "启用原价；另有停用更低进价（未进阶梯）"
            elif is_yuan:
                judge = "对外原价"
            elif disabled_cheap:
                judge = "有公开阶梯；另见停用更低进价"
            else:
                judge = "有公开阶梯"

            z3 = list(zhes) + [None] * 3
            vals = [
                m["vendor"],
                mid,
                m["display"],
                fam if fam is not None else "—",
                fmt_tier_zhe(z3[0]),
                fmt_tier_zhe(z3[1]),
                fmt_tier_zhe(z3[2]),
                src,
                f"{min_en:.2f}" if min_en is not None else "—",
                en_routes or "—",
                dis_txt or "—",
                judge,
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(rr, c, v)
                cell.font = mono if c in (2, 4, 5, 6, 7, 9) else body
                if is_yuan:
                    cell.fill = fill_yuan
                elif min_en is not None and min_en < 0.78:
                    cell.fill = fill_ok
            rr += 1
        return rr

    r = emit_resolution_rows(r0 + 2, models)

    # ---- 二-B：生图全量 ----
    img_models = []
    seen_img: set[str] = set()
    for m in image_rows:
        if m["model_id"] in seen_img:
            continue
        seen_img.add(m["model_id"])
        img_models.append(m)
    r += 1
    ws.cell(r, 1, "二-B、生图全量 · 报价解析").font = sec_font
    ws.cell(r, 1).fill = fill_sec
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    r += 1
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h)
        cell.font = bold
        cell.fill = fill_head
    r = emit_resolution_rows(r + 1, img_models)

    # ---- 二-C：生视频全量 ----
    vid_models = []
    seen_vid: set[str] = set()
    for m in video_rows:
        if m["model_id"] in seen_vid:
            continue
        seen_vid.add(m["model_id"])
        vid_models.append(m)
    r += 1
    ws.cell(r, 1, "二-C、生视频全量 · 报价解析").font = sec_font
    ws.cell(r, 1).fill = fill_sec
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    r += 1
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h)
        cell.font = bold
        cell.fill = fill_head
    r = emit_resolution_rows(r + 1, vid_models)

    # 原价名单扩及生图/生视频
    for m in img_models + vid_models:
        fam = resolve_family(m["model_id"], model_families, recommended)
        zhes = public_tiers_for(m["model_id"], fam)
        if zhes and all(z is not None and z >= 9.999 for z in zhes):
            if m["model_id"] not in yuanjia_models:
                yuanjia_models.append(m["model_id"])

    # ---- 第三节：原价专项 + 停用明细 ----
    r += 1
    ws.cell(
        r, 1, "三、对外原价专项 · 启用=1.0 与停用更低进价明细（生文+生图+生视频）"
    ).font = sec_font
    ws.cell(r, 1).fill = fill_sec
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    r += 1
    ws.cell(
        r,
        1,
        "说明：下表逐条列出原价模型的启用线路（成本比）及全部非启用且成本比<0.97 的线路，"
        "作为「为何对外原价 / 是否可改阶梯」的报价依据。",
    ).font = hint
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    r += 1
    h2 = [
        "模型 ID",
        "启停",
        "折扣列",
        "成本/官方比",
        "线路名",
        "源文件",
        "对报价含义",
    ]
    for c, h in enumerate(h2, 1):
        cell = ws.cell(r, c, h)
        cell.font = bold
        cell.fill = fill_head
    r += 1

    for mid in yuanjia_models:
        lines = by_mid.get(mid, [])
        if not lines:
            ws.cell(r, 1, mid).font = mono
            ws.cell(r, 7, "线路源无记录").font = body
            r += 1
            continue
        # 启用在前，再停用低价，再其它
        def sort_key(x):
            en = 0 if x["status"] == "启用" else 1
            ratio = x["min_ratio"] if x["min_ratio"] is not None else 9
            return (en, ratio, x["source"], x["route"])

        for x in sorted(lines, key=sort_key):
            # 原价专项：启用全记；非启用只记更低进价
            if x["status"] != "启用" and (
                x["min_ratio"] is None or x["min_ratio"] >= 0.97
            ):
                continue
            if x["status"] == "启用":
                meaning = "当前供给·进货≈挂牌 → 支撑对外原价"
            else:
                meaning = "停用更低进价·未进商务归属/对外阶梯；若启用需回灌成本族"
            vals = [
                mid,
                x["status"] or "—",
                str(x["discount"]) if x["discount"] is not None else "—",
                f"{x['min_ratio']:.3f}" if x["min_ratio"] is not None else "—",
                x["route"] or "—",
                x["source"],
                meaning,
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v)
                cell.font = mono if c in (1, 4) else body
                if x["status"] != "启用":
                    cell.fill = fill_yuan
            r += 1

    r += 1
    ws.cell(r, 1, "四、口径备忘").font = sec_font
    ws.cell(r, 1).fill = fill_sec
    r += 1
    for line in [
        "1) 对外三档（$5k/$10k/$50k）取公开定折矩阵；对内商务仍五档；不抄 L3b 全列深折。",
        "2) 模型→成本族：商务归属 + 交叉推荐（厚利优先）。",
        "3) 1.0 族 / 未映射 → 对外原价；0.78 → 浅折；Claude 非 1.0 → 对齐 opus 公开阶梯。",
        "4) 停用线路不改变对外价；重新启用并回灌商务表后，需重跑本脚本刷新 01 与外发表。",
        "5) ≤0.50：0.40 最深 5.5、0.50 最深 6.0；对客折底线暂定 5.5；097 全停用仅线路扫描。",
    ]:
        ws.cell(r, 1, line).font = body
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        r += 1

    widths = [14, 28, 22, 12, 10, 10, 10, 22, 12, 36, 40, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{r0 + 2}"
    print(
        f"price resolution: text={len(models)} image={len(img_models)} "
        f"video={len(vid_models)} yuanjia={len(yuanjia_models)} "
        f"dormant_cheaper_rows={len(dormant_cheaper)} route_rows={len(route_rows)}"
    )


def write_l3a_internal_to_commercial(
    meta: list[tuple[str, str]],
    models: list[dict],
    image_rows: list[dict],
    video_rows: list[dict],
    model_families: dict[str, list[float]],
    recommended: dict[str, float],
) -> None:
    """回写商务总册 01_报价解析汇总（勿外发）；不再写 90/91/src。"""
    path = SRC_COMMERCIAL
    if not path.exists():
        print(f"skip L3a internal sheets: missing {path}")
        return
    wb = openpyxl.load_workbook(path)
    # 清理旧页签（迁移期）
    for name in (
        "90_L3a对外说明",
        "91_L3a映射索引",
        "92_报价解析汇总",
        "00_说明_L3a",
        "99_对内索引_勿外发",
        "01_商务洽谈总表",
        "02_交叉模型",
    ):
        if name in wb.sheetnames:
            del wb[name]
    # 清理旧 src_* 
    for name in list(wb.sheetnames):
        if name.startswith("src_"):
            del wb[name]

    write_price_resolution_sheet(
        wb, meta, models, image_rows, video_rows, model_families, recommended
    )

    # 刷新 00_说明 · 解析回写日
    if "00_说明" in wb.sheetnames:
        wsr = wb["00_说明"]
        found = False
        for r in range(1, wsr.max_row + 1):
            if str(wsr.cell(r, 1).value or "").strip() == "01最近回写":
                wsr.cell(r, 2, date.today().isoformat())
                found = True
                break
        if not found:
            r = wsr.max_row + 1
            wsr.cell(r, 1, "01最近回写").font = Font(name="PingFang SC", bold=True)
            wsr.cell(r, 2, date.today().isoformat()).font = Font(name="PingFang SC")

    wb.save(path)
    print(f"updated {SHEET_RESOLUTION} → {path.name} sheets={wb.sheetnames}")


def build():
    models = load_models()
    image_rows = load_image_rows()
    video_rows = load_video_rows()
    _family_tiers_l3b, model_families, recommended = load_commercial()
    image_rows = sort_image_rows_discount_first(
        image_rows, model_families, recommended
    )
    video_rows = sort_video_rows_discount_first(
        video_rows, model_families, recommended
    )
    today = date.today()
    valid_until = "2026-09-30 24:00"
    st = _sheet_styles()
    thin = st["thin"]
    header_fill = st["header_fill"]
    alt = st["alt"]
    body_font = st["body_font"]
    num_font = st["num_font"]
    white_bold = st["white_bold"]
    center = st["center"]
    left = st["left"]

    with_ladder = 0
    for m in models:
        fam = resolve_family(m["model_id"], model_families, recommended)
        zhes = public_tiers_for(m["model_id"], fam)
        if zhes and any(z is not None and z < 9.999 for z in zhes):
            with_ladder += 1
    image_models = len({r["model_id"] for r in image_rows})
    video_models = len({r["model_id"] for r in video_rows})
    meta = [
        ("外发文件", OUT.name),
        ("定位", "L3a 对外标准档报价单（整本可发；无对内 Sheet）"),
        (
            "形态",
            "00_折扣一览（仅有折扣）· 01_生文 · 02_生图 · 03_生视频；"
            "刊例价 + 用量档对客折（对外三档）",
        ),
        ("不含", "上游成本折/线路/中转站/GM；报价依据在商务总册 01_报价解析汇总"),
        (
            "刊例价来源",
            "L2 线上刊例 · prices-api-text/image/video.json（与 /v1/prices 一致）；"
            "生文本轮上浮 hy-mt2-plus、glm-5.2（刊例策略 §3.1）",
        ),
        (
            "小数规则",
            "对外金额统一两位小数；原值第 3 位及以后有数则向上进位到分",
        ),
        (
            "客户档",
            "对外三档：≥$5k · ≥$10k · ≥$50k（表头标「对外」；对内商务仍五档）",
        ),
        (
            "阶梯来源",
            "定价方案-v0 公开定折矩阵；模型→成本族取自本商务表；"
            "≤0.50：0.40 最深 5.5、0.50 最深 6.0；对客折底线暂定 5.5；"
            "图/视频多数 1.0 则写原价；不抄 L3b 商务全五档",
        ),
        (
            "公开定折摘要",
            "顺序 对外$5k/$10k/$50k · "
            "0.40: 6.2/6.0/5.5；0.50: 6.8/6.5/6.0；"
            "0.60: 8.2/7.6/6.7；0.65: 8.5/8.2/7.2；"
            "0.70: 9.2/8.8/7.8；0.75: 9.0/8.7/8.2；"
            "0.85: 9.8/9.7/9.4；≥0.78/未映射: 原价（Claude/单模特例除外）",
        ),
        ("达档口径", "企业户累积消耗（按刊例价计）"),
        ("生成日期", today.isoformat()),
        ("有效期（初版暂定）", f"至 {valid_until}；过期请向运营索取新版"),
        (
            "单位",
            "生文：USD/百万 tokens；生图：USD/张或百万 tokens；"
            "生视频：USD/秒或百万 video tokens（见规格列）",
        ),
        (
            "覆盖",
            f"生文 {len(models)} 款（公开阶梯非原价约 {with_ladder}）；"
            f"生图 {image_models} 款 / {len(image_rows)} 规格行；"
            f"生视频 {video_models} 款 / {len(video_rows)} 规格行",
        ),
        ("重建", "python3 pricing/scripts/build_outward_quote_standard.py"),
    ]

    # ---- 对外 workbook：仅客户可发 Sheet ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "01_生文"
    # A厂商 B模型ID C显示名 D目录入 E目录出 F目录缓 | 各用量档（仅对客折，入出同折）
    n_cols = 6 + len(TIERS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws["A1"] = "Trinity · 生文标准档报价"
    ws["A1"].font = Font(name="PingFang SC", size=18, bold=True, color="1D2939")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws["A2"] = (
        f"【生文】单位：USD / 百万 tokens　|　有效期：{valid_until}，过期作废，请联系商务获取最新报价。"
        "　阶梯达档：企业户累积消耗（按刊例价计）；更大用量或合同价请联系商务。"
        "　用量档列仅列对客折（入/出同折）；折后价 = 刊例价 × 折数。"
    )
    ws["A2"].font = Font(name="PingFang SC", size=10, color="475467")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 40

    # 单行表头（用量档不再分子列「输入/输出」）
    headers = [
        "厂商",
        "模型 ID",
        "显示名",
        "刊例价·输入",
        "刊例价·输出",
        "刊例价·缓存",
    ] + [label for _, label, _ in TIERS]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(4, i, h)
        style_cell(cell, fill=header_fill, font=white_bold, align=center, border=thin)
    ws.row_dimensions[4].height = 24

    prev_vendor = None
    for idx, m in enumerate(models):
        r = 5 + idx
        show_vendor = m["vendor"] if m["vendor"] != prev_vendor else ""
        prev_vendor = m["vendor"]
        fam = resolve_family(m["model_id"], model_families, recommended)
        zhes = public_tiers_for(m["model_id"], fam)

        vals = [
            show_vendor,
            m["model_id"],
            m["display"],
            fmt_num(m["inp"]),
            fmt_num(m["out"]),
            fmt_cache(m.get("cache")),
        ]
        for ti, (_, _, _) in enumerate(TIERS):
            if not zhes or zhes[ti] is None:
                vals.append("—")
            else:
                vals.append(fmt_tier_zhe(zhes[ti]))

        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            style_cell(
                cell,
                font=num_font if c >= 4 else body_font,
                align=center if c == 1 or c >= 4 else left,
                border=thin,
                fill=alt if idx % 2 else None,
            )

    widths = [12, 26, 26, 12, 12, 12] + [14] * len(TIERS)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "G5"

    # ---- 02_生图 ----
    ws_img = wb.create_sheet("02_生图", 1)
    n_img = 6 + len(TIERS)
    ws_img.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_img)
    ws_img["A1"] = "Trinity · 生图标准档报价"
    ws_img["A1"].font = Font(name="PingFang SC", size=18, bold=True, color="1D2939")
    ws_img.row_dimensions[1].height = 28
    ws_img.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_img)
    ws_img["A2"] = (
        f"【生图】有效期：{valid_until}，过期作废，请联系商务获取最新报价。"
        "　阶梯达档：企业户累积消耗（按刊例价计）。"
        "　规格多为分辨率（USD/张）；token 生图见单位列。"
        "　用量档仅列对客折；折后价 = 刊例价 × 折数。"
        "　刊例价 = 线上 L2（/v1/prices），对外金额两位小数进位。"
    )
    ws_img["A2"].font = Font(name="PingFang SC", size=10, color="475467")
    ws_img["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws_img.row_dimensions[2].height = 44

    img_headers = [
        "厂商",
        "模型 ID",
        "显示名",
        "规格",
        "单位",
        "刊例价",
    ] + [label for _, label, _ in TIERS]
    for i, h in enumerate(img_headers, 1):
        cell = ws_img.cell(4, i, h)
        style_cell(cell, fill=header_fill, font=white_bold, align=center, border=thin)
    ws_img.row_dimensions[4].height = 24

    prev_vendor = None
    for idx, m in enumerate(image_rows):
        r = 5 + idx
        show_vendor = m["vendor"] if m["vendor"] != prev_vendor else ""
        prev_vendor = m["vendor"]
        fam = resolve_family(m["model_id"], model_families, recommended)
        zhes = public_tiers_for(m["model_id"], fam)
        cat_cell = (
            m["catalog_display"]
            if m.get("catalog_display")
            else fmt_num(m["catalog"])
        )
        vals = [
            show_vendor,
            m["model_id"],
            m["display"],
            m["spec"],
            m["unit"],
            cat_cell,
        ]
        for ti, (_, _, _) in enumerate(TIERS):
            if not zhes or zhes[ti] is None:
                vals.append("—")
            else:
                vals.append(fmt_tier_zhe(zhes[ti]))
        for c, v in enumerate(vals, 1):
            cell = ws_img.cell(r, c, v)
            style_cell(
                cell,
                font=num_font if c >= 4 else body_font,
                align=center if c == 1 or c >= 4 else left,
                border=thin,
                fill=alt if idx % 2 else None,
            )

    for i, w in enumerate([12, 22, 22, 10, 16, 28] + [14] * len(TIERS), 1):
        ws_img.column_dimensions[get_column_letter(i)].width = w
    ws_img.freeze_panes = "G5"

    # ---- 03_生视频 ----
    ws_vid = wb.create_sheet("03_生视频", 2)
    n_vid = 6 + len(TIERS)
    ws_vid.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_vid)
    ws_vid["A1"] = "Trinity · 生视频标准档报价"
    ws_vid["A1"].font = Font(name="PingFang SC", size=18, bold=True, color="1D2939")
    ws_vid.row_dimensions[1].height = 28
    ws_vid.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_vid)
    ws_vid["A2"] = (
        f"【生视频】有效期：{valid_until}，过期作废，请联系商务获取最新报价。"
        "　阶梯达档：企业户累积消耗（按刊例价计）。"
        "　规格多为分辨率或时长计费（USD/秒）；token 生视频见单位列。"
        "　用量档仅列对客折；折后价 = 刊例价 × 折数。"
        "　刊例价 = 线上 L2（/v1/prices），对外金额两位小数进位。"
    )
    ws_vid["A2"].font = Font(name="PingFang SC", size=10, color="475467")
    ws_vid["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws_vid.row_dimensions[2].height = 44

    vid_headers = [
        "厂商",
        "模型 ID",
        "显示名",
        "规格",
        "单位",
        "刊例价",
    ] + [label for _, label, _ in TIERS]
    for i, h in enumerate(vid_headers, 1):
        cell = ws_vid.cell(4, i, h)
        style_cell(cell, fill=header_fill, font=white_bold, align=center, border=thin)
    ws_vid.row_dimensions[4].height = 24

    prev_vendor = None
    for idx, m in enumerate(video_rows):
        r = 5 + idx
        show_vendor = m["vendor"] if m["vendor"] != prev_vendor else ""
        prev_vendor = m["vendor"]
        fam = resolve_family(m["model_id"], model_families, recommended)
        zhes = public_tiers_for(m["model_id"], fam)
        cat_cell = (
            m["catalog_display"]
            if m.get("catalog_display")
            else fmt_num(m["catalog"])
        )
        vals = [
            show_vendor,
            m["model_id"],
            m["display"],
            m["spec"],
            m["unit"],
            cat_cell,
        ]
        for ti, (_, _, _) in enumerate(TIERS):
            if not zhes or zhes[ti] is None:
                vals.append("—")
            else:
                vals.append(fmt_tier_zhe(zhes[ti]))
        for c, v in enumerate(vals, 1):
            cell = ws_vid.cell(r, c, v)
            style_cell(
                cell,
                font=num_font if c >= 4 else body_font,
                align=center if c == 1 or c >= 4 else left,
                border=thin,
                fill=alt if idx % 2 else None,
            )

    for i, w in enumerate([12, 26, 22, 14, 22, 14] + [14] * len(TIERS), 1):
        ws_vid.column_dimensions[get_column_letter(i)].width = w
    ws_vid.freeze_panes = "G5"

    discount_rows = collect_discounted_models(
        models, image_rows, video_rows, model_families, recommended
    )
    write_discount_overview_sheet(
        wb,
        discount_rows,
        valid_until=valid_until,
        thin=thin,
        white_bold=white_bold,
        body_font=body_font,
        num_font=num_font,
        center=center,
        left=left,
    )

    wb.save(OUT)
    patch_xlsx_for_wechat(OUT)  # 微信预览：补 applyFill 等
    mapped = sum(
        1
        for m in models
        if resolve_family(m["model_id"], model_families, recommended) is not None
    )
    print(
        f"wrote {OUT} sheets={wb.sheetnames} text={len(models)} "
        f"mapped_family={mapped} public_ladder≈{with_ladder} "
        f"image_models={image_models} image_rows={len(image_rows)} "
        f"video_models={video_models} video_rows={len(video_rows)}"
    )

    # 说明 + 映射索引 → 商务内部表（对内）；外发仅 OUT 一份
    write_l3a_internal_to_commercial(
        meta, models, image_rows, video_rows, model_families, recommended
    )
    if SRC_COMMERCIAL.exists():
        patch_xlsx_for_wechat(SRC_COMMERCIAL)


if __name__ == "__main__":
    build()
