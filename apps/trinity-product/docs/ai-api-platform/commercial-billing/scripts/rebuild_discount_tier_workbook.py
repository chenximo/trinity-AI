#!/usr/bin/env python3
"""Rebuild 商务洽谈折扣总表.xlsx from 线路管理 exports.

SOP: ../discount-tier-workbook-sop.md
折数真源: pricing-strategy-evidence-chain.md — change tiers there first, then FAMILY_TIERS here.
"""

from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_PRICING_SCRIPTS = Path(__file__).resolve().parents[6] / "pricing" / "scripts"
if str(_PRICING_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_PRICING_SCRIPTS))
from xlsx_wechat_compat import patch_xlsx_for_wechat  # noqa: E402

# ---------------------------------------------------------------------------
# Config — register new cost-family exports here
# ---------------------------------------------------------------------------

PRICING_ROOT = Path(__file__).resolve().parents[6] / "pricing"
PRICING_INPUT = PRICING_ROOT / "input"
ROUTES_TEXT = PRICING_INPUT / "routes-20260809-text"
ROUTES_IMAGE = PRICING_INPUT / "routes-20260809-image"
ROUTES_VIDEO = PRICING_INPUT / "routes-20260809-video"
# 一本总册（8 Sheet）；线路整表不进册，只留 input 归档
OUT = PRICING_ROOT / "output" / "商务洽谈折扣总表.xlsx"
OUTWARD_FILE = "Trinity模型报价表.xlsx"

SHEET_00 = "00_说明"
SHEET_01 = "01_报价解析汇总"
SHEET_10 = "10_商务总表-生文"
SHEET_11 = "11_交叉模型-生文"
SHEET_20 = "20_商务总表-生图"
SHEET_21 = "21_交叉模型-生图"
SHEET_30 = "30_商务总表-生视频"
SHEET_31 = "31_交叉模型-生视频"

# (export path, 折扣列原文, d_cost key)
# 生文：Downloads (1)～(7) → routes-20260809-text/
SOURCES_TEXT: list[tuple[Path, str, str]] = [
    (ROUTES_TEXT / "050.xlsx", "5折", "0.50"),
    (ROUTES_TEXT / "060.xlsx", "6折", "0.60"),
    (ROUTES_TEXT / "065.xlsx", "6.5折", "0.65"),
    (ROUTES_TEXT / "070.xlsx", "7折", "0.70"),
    (ROUTES_TEXT / "075.xlsx", "7.5折", "0.75"),
    (ROUTES_TEXT / "078.xlsx", "7.8折", "0.78"),
    (ROUTES_TEXT / "100.xlsx", "原价", "1.0"),
]
# 生图：Downloads (8)～(10) → routes-20260809-image/
SOURCES_IMAGE: list[tuple[Path, str, str]] = [
    (ROUTES_IMAGE / "065.xlsx", "6.5折", "0.65"),
    (ROUTES_IMAGE / "075.xlsx", "7.5折", "0.75"),
    (ROUTES_IMAGE / "100.xlsx", "原价", "1.0"),
]
# 生视频：Downloads (11)～(15)(17)(18) → routes-20260809-video/
SOURCES_VIDEO: list[tuple[Path, str, str]] = [
    (ROUTES_VIDEO / "040.xlsx", "4折", "0.40"),
    (ROUTES_VIDEO / "060.xlsx", "6折", "0.60"),
    (ROUTES_VIDEO / "070.xlsx", "7折", "0.70"),
    (ROUTES_VIDEO / "075.xlsx", "7.5折", "0.75"),
    (ROUTES_VIDEO / "085.xlsx", "8.5折", "0.85"),
    (ROUTES_VIDEO / "097.xlsx", "9.7折", "0.97"),
    (ROUTES_VIDEO / "100.xlsx", "原价", "1.0"),
]

# d_cost -> label, tiers[Plus..Enterprise], band, public_flag, no_ladder
# tiers values: "原价" | "9.0" | "—"
# 行序：上游成本折由低→高（0.40…1.0）；档内对客折浅→深。中间锚点暂不硬定。
# 0.40 / 0.50：最深档已拍（0.40→5.5 · 0.50→6.0）；报价表对客折底线暂定 5.5
FAMILY_TIERS: list[tuple] = [
    (0.40, "0.40（4折）", ["6.5", "6.2", "6.0", "5.8", "5.5"],
     "最深5.5（表底线；生视频厚利）", "是（对外五档）", False),
    (0.50, "0.50（5折）", ["7.0", "6.8", "6.5", "6.2", "6.0"],
     "最深6.0（不低于表底线5.5）", "是（对外五档）", False),
    (0.60, "0.60（6折）", ["8.5", "8.2", "7.6", "7.0", "6.7"], "—", "待定", False),
    (0.65, "0.65（6.5折）·主锚", ["9.0", "8.5", "8.2", "7.6", "7.2"],
     "Growth 8.5～8.2；Scale 7.5～7.7", "否（公开另文）", False),
    (0.70, "0.70（7折）", ["9.7", "9.2", "8.8", "8.2", "7.8"], "—", "待定", False),
    (0.75, "0.75（7.5折）", ["原价", "9.8", "9.4", "8.8", "8.3"], "—", "待定", False),
    (0.78, "0.78（7.8折）", ["原价", "原价", "9.8", "9.1", "8.7"], "—", "待定", False),
    (0.80, "0.80（8折·薄利）", ["9.8", "9.5", "9.3", "9.1", "8.9"], "—", "否·仅商务", False),
    (0.85, "0.85（8.5折·薄利）", ["9.9", "9.8", "9.7", "9.5", "9.4"], "—", "否·仅商务", False),
    (0.90, "0.90（9折·薄利）", ["原价", "9.9", "9.9", "9.8", "9.7"], "—", "否·仅商务", False),
    (0.97, "0.97（9.7折）·不设阶梯", ["—", "—", "—", "—", "—"],
     "不设用量阶梯；对客原价（刊例 GM≈3%）", "否·仅刊例/商务点名", True),
    (1.0, "1.0（原价）·不设阶梯", ["—", "—", "—", "—", "—"],
     "不设用量阶梯；进货≈挂牌；对客刊例（GM≈0%）", "否·仅刊例/商务点名", True),
]

FAMILY_ORDER = [
    "0.40", "0.50", "0.60", "0.65", "0.70", "0.75", "0.78",
    "0.80", "0.85", "0.90", "0.97", "1.0",
]
CROSS_ROUTE_COLS = [
    "0.40", "0.50", "0.60", "0.65", "0.70", "0.75", "0.78", "0.85", "1.0",
]

TIER_HEADERS = [
    "上游成本折",
    "≥$1k（对内·Plus）",
    "≥$5k（对内·Mid）",
    "≥$10k（对内·Growth）",
    "≥$30k（对内·Scale）",
    "≥$50k（对内·Enterprise）",
    "操作带",
    "模型数",
    "模型（ID※：主线路）",
    "交叉（跨成本折）",
    "是否进公开阶梯",
]

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

THIN = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)
def _solid(rgb6: str) -> PatternFill:
    """不透明实心底色（FF 前缀，避免微信预览把 00alpha 当成透明）。"""
    h = str(rgb6).strip().lstrip("#").upper()
    if len(h) == 8:
        h = "FF" + h[2:]
    elif len(h) == 6:
        h = "FF" + h
    return PatternFill(fill_type="solid", fgColor=h, bgColor=h)


FILL_HINT = _solid("FEF9C3")
FILL_HEAD = _solid("F1F5F9")
FILL_ANCHOR = _solid("FFF7ED")
FILL_FILLED = _solid("F0FDF4")
FILL_NOLADDER = _solid("F1F5F9")
FILL_ROW = _solid("FFFFFF")
FILL_ALT = _solid("F8FAFC")
FILL_TITLE = _solid("FEE2E2")
FILL_REC = _solid("DCFCE7")
FILL_CROSS = _solid("FEF2F2")  # 交叉列浅底

# 模型格富文本：※ 用红色加粗，其余保持深灰等宽
INLINE_MODEL = InlineFont(rFont="Menlo", sz=9, color="334155")
INLINE_STAR = InlineFont(rFont="Menlo", sz=9, color="B91C1C", b=True)
HEAD_BORDER = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="medium", color="334155"),
)


def font(bold=False, size=10, color="1E293B", name="PingFang SC"):
    return Font(name=name, bold=bold, size=size, color=color)


# ---------------------------------------------------------------------------
# Load
# ---------------------------------------------------------------------------

def fix_route(route: str) -> str:
    if not route:
        return ""
    return route.replace("??????", "网聚云联")


def discount_matches(cell, label: str) -> bool:
    """Match 折扣列：原文「6.5折」或原价族的 1 / 1.0 /「原价」."""
    if cell is None:
        return False
    if label == "原价":
        if cell in (1, 1.0, "1", "1.0", "原价", "10折", "十折"):
            return True
        try:
            return abs(float(cell) - 1.0) < 1e-9
        except (TypeError, ValueError):
            return str(cell).strip() in ("原价", "10折", "十折")
    return str(cell).strip() == str(label).strip()


def parse_cost_ratios_from_rows(rows_values) -> dict[str, float]:
    """From 线路管理 rows, min(成本/官方) per model (forward-fill model code)."""
    import re

    ratios: dict[str, list[float]] = defaultdict(list)
    cur = None
    for row in rows_values:
        if not row:
            continue
        code = row[0]
        if code:
            cur = str(code).strip()
        if not cur:
            continue
        cost_cell = row[11] if len(row) > 11 else None
        if not cost_cell:
            continue
        for m in re.finditer(r"官方\s*([\d.]+)\s*/\s*成本\s*([\d.]+)", str(cost_cell)):
            o, c = float(m.group(1)), float(m.group(2))
            if o > 0:
                ratios[cur].append(c / o)
    return {mid: min(vs) for mid, vs in ratios.items() if vs}


def load_source(path: Path, discount_label: str):
    """Return (pairs for summary, route map for cross, cost_ratio_by_mid)."""
    if not path.exists():
        raise FileNotFoundError(f"SOURCE missing: {path}")
    wb = load_workbook(path, data_only=True)
    ws = wb["线路管理"]
    all_rows = list(ws.iter_rows(values_only=True))
    cost_ratio = parse_cost_ratios_from_rows(all_rows)
    pairs = []  # (id, route) for model cell
    routes = defaultdict(list)  # id -> [(route, pri, weight)]
    for i, row in enumerate(all_rows, 1):
        if i == 1 or not row or not row[0]:
            continue
        if not discount_matches(row[8], discount_label) or row[10] != "启用":
            continue
        mid = row[0]
        route = fix_route(row[3] or "")
        pri = str(row[4] or "—")
        weight = str(row[5] or "—")
        pairs.append((mid, route))
        entry = (route, pri, weight)
        if entry not in routes[mid]:
            routes[mid].append(entry)
    return pairs, routes, cost_ratio


def gm_pct(d_cost: float, tier: str):
    if tier in ("—", "-"):
        return None
    if tier == "原价":
        return (1 - d_cost) * 100
    return (1 - d_cost / (float(tier) / 10.0)) * 100


def fmt_gm(gm: float) -> str:
    half = round(gm * 2) / 2
    if abs(gm - half) < 0.35:
        return f"{int(half)}" if half == int(half) else f"{half:.1f}"
    return f"{gm:.1f}"


def gm_label(d_cost: float, tier: str) -> str:
    if tier in ("—", "-"):
        return "—"
    if tier == "原价":
        return f"原价（GM {fmt_gm(gm_pct(d_cost, tier))}%）"
    return f"{tier}（GM {fmt_gm(gm_pct(d_cost, tier))}%）"


def models_cell(
    pairs: list[tuple[str, str]],
    cross_ids: set[str] | None = None,
):
    if not pairs:
        return "（待补）"
    cross_ids = cross_ids or set()
    has_cross = any(mid in cross_ids for mid, _ in pairs)
    if not has_cross:
        lines = []
        for i, (mid, route) in enumerate(pairs):
            suffix = "；" if i < len(pairs) - 1 else ""
            lines.append(f"{mid}：{route}{suffix}")
        return "\n".join(lines)

    blocks: list[TextBlock] = []
    for i, (mid, route) in enumerate(pairs):
        suffix = "；" if i < len(pairs) - 1 else ""
        eol = "\n" if i < len(pairs) - 1 else ""
        blocks.append(TextBlock(INLINE_MODEL, mid))
        if mid in cross_ids:
            blocks.append(TextBlock(INLINE_STAR, "※"))
        blocks.append(TextBlock(INLINE_MODEL, f"：{route}{suffix}{eol}"))
    return CellRichText(*blocks)


def cross_cell(
    pairs: list[tuple[str, str]],
    all_routes: dict[str, dict[str, list]],
) -> str:
    marks = []
    for mid, _ in pairs:
        fam_map = all_routes.get(mid) or {}
        if len(fam_map) < 2:
            continue
        fams = sorted(
            fam_map.keys(),
            key=lambda x: FAMILY_ORDER.index(x) if x in FAMILY_ORDER else 99,
        )
        marks.append(f"{mid}→{'·'.join(fams)}")
    if not marks:
        return "—"
    return f"※{len(marks)} 见交叉表｜" + "；".join(marks)


def row_height_for(n: int) -> float:
    if n <= 0:
        return 38
    return max(42, n * 14.5 + 16)


def write_placeholder_sheet(wb: Workbook, title: str, modality: str, kind: str):
    ws = wb.create_sheet(title)
    ws["A1"] = f"{title} · 待补"
    ws["A1"].font = font(bold=True, size=12, color="B91C1C")
    ws["A2"] = (
        f"{modality} · {kind}尚未回灌。"
        f"有线路导出后登记到脚本 SOURCES_* 并重跑 rebuild。"
        f"整表原料归档在 pricing/input/，不进本总册。"
    )
    ws["A2"].font = font(size=10, color="64748B")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 40
    ws.column_dimensions["A"].width = 88


def load_modality(sources: list[tuple[Path, str, str]], tag: str):
    loaded = {}
    cost_ratios_by_source: dict[str, dict[str, float]] = {}
    for path, label, key in sources:
        pairs, routes, cost_ratio = load_source(path, label)
        loaded[key] = (pairs, routes, path.name, label)
        cost_ratios_by_source[key] = cost_ratio
        print(f"[{tag}] loaded {key}: {len(pairs)} models from {path.name}")

    if "1.0" in loaded and "0.78" in loaded:
        pairs100, routes100, fname100, label100 = loaded["1.0"]
        pairs078, routes078, fname078, label078 = loaded["0.78"]
        ratios = cost_ratios_by_source.get("1.0", {})
        keep100, move078 = [], []
        seen_move = set()
        for mid, route in pairs100:
            r = ratios.get(mid)
            if r is not None and abs(r - 0.78) < 0.03 and mid not in seen_move:
                move078.append((mid, route))
                seen_move.add(mid)
                for e in routes100.get(mid, []):
                    if e not in routes078[mid]:
                        routes078[mid].append(e)
            else:
                keep100.append((mid, route))
        for mid in seen_move:
            routes100.pop(mid, None)
        have078 = {m for m, _ in pairs078}
        for mid, route in move078:
            if mid not in have078:
                pairs078.append((mid, route))
                have078.add(mid)
        loaded["1.0"] = (keep100, routes100, fname100, label100)
        loaded["0.78"] = (pairs078, routes078, fname078, label078)
        if move078:
            print(
                f"[{tag}] reassigned {len(move078)} from 1.0→0.78: "
                + ", ".join(m for m, _ in move078)
            )

    if "1.0" in loaded:
        thicker_ids: set[str] = set()
        for k, (pairs, *_rest) in loaded.items():
            if k == "1.0":
                continue
            thicker_ids.update(mid for mid, _ in pairs)
        pairs100, routes100, fname100, label100 = loaded["1.0"]
        deduped = [(m, r) for m, r in pairs100 if m not in thicker_ids]
        dropped = len(pairs100) - len({m for m, _ in deduped})
        routes100 = {m: v for m, v in routes100.items() if m not in thicker_ids}
        loaded["1.0"] = (deduped, routes100, fname100, label100)
        print(f"[{tag}] 1.0 after exclude thicker: {len(deduped)} (dropped {dropped})")

    all_routes: dict[str, dict[str, list]] = defaultdict(lambda: defaultdict(list))
    for key, (_pairs, routes, *_rest) in loaded.items():
        for mid, ents in routes.items():
            for e in ents:
                if e not in all_routes[mid][key]:
                    all_routes[mid][key].append(e)

    pairs_by_key = {k: v[0] for k, v in loaded.items()}
    cross_id_set = {mid for mid, m in all_routes.items() if len(m) >= 2}
    return loaded, pairs_by_key, all_routes, cross_id_set


KEY_MAP = {
    0.40: "0.40", 0.50: "0.50", 0.65: "0.65", 0.60: "0.60", 0.70: "0.70",
    0.75: "0.75", 0.78: "0.78", 0.80: "0.80", 0.85: "0.85", 0.90: "0.90",
    0.97: "0.97", 1.0: "1.0",
}


def write_main_sheet(
    wb: Workbook,
    title: str,
    pairs_by_key: dict,
    all_routes: dict,
    cross_id_set: set[str],
    cross_sheet_name: str,
):
    ws = wb.create_sheet(title)
    for col, h in enumerate(TIER_HEADERS, 1):
        cell = ws.cell(1, col, h)
        cell.font = font(bold=True)
        cell.fill = FILL_HEAD
        cell.border = HEAD_BORDER
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    for i, (d_cost, label, tiers, band, public, noladder) in enumerate(FAMILY_TIERS):
        key = KEY_MAP[d_cost]
        pairs = pairs_by_key.get(key, [])
        r = 2 + i
        vals = [
            label,
            *[gm_label(d_cost, t) for t in tiers],
            band,
            len(pairs) if pairs else "—",
            models_cell(pairs, cross_id_set),
            cross_cell(pairs, all_routes),
            public,
        ]
        if d_cost in (0.65, 0.40):
            fill = FILL_ANCHOR
        elif noladder and pairs:
            fill = FILL_NOLADDER
        elif pairs:
            fill = FILL_FILLED
        else:
            fill = FILL_ALT if i % 2 else FILL_ROW
        for col, v in enumerate(vals, 1):
            cell = ws.cell(r, col, v)
            cell.border = THIN
            cell.fill = fill
            if col == 1:
                cell.font = font(bold=True)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            elif 2 <= col <= 6:
                cell.font = font()
                cell.alignment = Alignment(
                    vertical="top", horizontal="center", wrap_text=True
                )
            elif col == 9:
                cell.font = Font(name="Menlo", size=9, color="334155")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            elif col == 10:
                cell.font = font(size=9, color="991B1B")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if v and v != "—":
                    cell.fill = FILL_CROSS
            elif col == 8:
                cell.font = font()
                cell.alignment = Alignment(vertical="top", horizontal="center")
            else:
                cell.font = font(size=9, color="64748B")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = row_height_for(len(pairs)) if pairs else 38

    note_row = 2 + len(FAMILY_TIERS) + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=11)
    c = ws.cell(
        note_row,
        1,
        "说明：行序=上游成本折低→高（0.40…1.0）｜档内对客折浅→深｜"
        "对内档 Plus→…→Enterprise｜门槛 $1k/$5k/$10k/$30k/$50k｜档位格=对客折（GM）｜"
        "达档=企业户累积消耗刊例·非预存｜Standard<$1k原价｜≥0.95及1.0原价族不设阶梯｜"
        f"模型 ID 后※=跨成本折 → 详查「{cross_sheet_name}」｜"
        "SOP 见 discount-tier-workbook-sop.md｜释义见证据链§3.0",
    )
    c.font = font(size=9, color="713F12")
    c.fill = FILL_HINT
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[note_row].height = 44
    for i, w in enumerate([18, 14, 14, 15, 15, 18, 28, 8, 48, 36, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"
    return ws


def write_cross_sheet(
    wb: Workbook,
    title: str,
    all_routes: dict,
    cross_id_set: set[str],
    main_sheet_name: str,
):
    cross_ids = sorted(cross_id_set)
    n_cross_cols = 2 + len(CROSS_ROUTE_COLS) + 2
    last_col = get_column_letter(n_cross_cols)
    ws2 = wb.create_sheet(title)
    ws2.merge_cells(f"A1:{last_col}1")
    c = ws2.cell(
        1,
        1,
        f"{title} · 宽表 · 共 {len(cross_ids)} 个（线路格含 优先级/权重）",
    )
    c.font = font(bold=True, size=11, color="991B1B")
    c.fill = FILL_TITLE
    c.alignment = Alignment(vertical="center")
    ws2.row_dimensions[1].height = 24

    ws2.merge_cells(f"A2:{last_col}2")
    c = ws2.cell(
        2,
        1,
        f"用法：点名先查本表 →「推荐成本折」→ 回 {main_sheet_name} 读阶梯折。"
        "线路格：线路名 · 优先级/权重。P1 多为默认主路。"
        "推荐：成本折越小越优先（0.50>0.60>0.65>…>1.0原价）。",
    )
    c.font = font(size=9, color="713F12")
    c.fill = FILL_HINT
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws2.row_dimensions[2].height = 48

    headers2 = [
        "Trinity ID",
        "涉及成本折",
        *[f"线路@{c}{'原价' if c == '1.0' else ''}" for c in CROSS_ROUTE_COLS],
        "推荐成本折",
        "商务提示",
    ]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(3, col, h)
        cell.font = font(bold=True)
        cell.fill = FILL_HEAD
        cell.border = THIN
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.row_dimensions[3].height = 28

    def recommend(families):
        for f in [
            "0.40", "0.50", "0.60", "0.65", "0.70", "0.75",
            "0.78", "0.85", "0.90", "0.97", "1.0",
        ]:
            if f in families:
                return f
        return sorted(families)[0]

    def tip(families, rec):
        fs = set(families)
        if "1.0" in fs and len(fs) > 1:
            return f"含原价线路；优先更厚利族 {rec}（勿用 1.0 作主谈折）"
        if fs >= {"0.65", "0.78"} and len(fs) == 2:
            return "厚利 0.65 优先；0.78 备选"
        if fs >= {"0.70", "0.75"}:
            return "双线：优先更厚利族；核默认路由后再定"
        return f"优先推荐成本折 {rec}；并结合格内 P1/权重看默认路由"

    def fmt_routes(entries):
        return "；".join(f"{route} · {pri}/{w}" for route, pri, w in entries)

    for i, mid in enumerate(cross_ids):
        r = 4 + i
        fam_map = all_routes[mid]
        families = sorted(
            fam_map.keys(),
            key=lambda x: FAMILY_ORDER.index(x) if x in FAMILY_ORDER else 99,
        )
        rec = recommend(families)
        route_cells = [
            fmt_routes(fam_map[f]) if f in fam_map else "" for f in CROSS_ROUTE_COLS
        ]
        vals = [mid, "、".join(families), *route_cells, rec, tip(families, rec)]
        rec_col = 2 + len(CROSS_ROUTE_COLS) + 1
        route_col_end = 2 + len(CROSS_ROUTE_COLS)
        fill = (
            FILL_ANCHOR
            if ("0.65" in fam_map or "0.50" in fam_map or "0.40" in fam_map)
            else FILL_ROW
        )
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(r, col, v)
            cell.border = THIN
            cell.fill = FILL_REC if col == rec_col else fill
            cell.font = (
                font(bold=True, color="166534")
                if col == rec_col
                else (
                    Font(name="Menlo", size=9, color="334155")
                    if col == 1 or 3 <= col <= route_col_end
                    else font(size=9)
                )
            )
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
                horizontal="center" if col in (2, rec_col) else "left",
            )
        ws2.row_dimensions[r].height = 40

    fr = 4 + max(len(cross_ids), 1) + 1
    if not cross_ids:
        ws2.cell(4, 1, "（本模态暂无跨成本折模型）").font = font(size=9, color="64748B")
    ws2.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=n_cross_cols)
    c = ws2.cell(
        fr,
        1,
        "附：同族多线路见 pricing/input 线路归档。"
        f"主表 {main_sheet_name} 不含优先级/权重；对外阶梯依据见 01_报价解析汇总。",
    )
    c.font = Font(name="PingFang SC", size=9, color="64748B", italic=True)
    c.alignment = Alignment(wrap_text=True)
    ws2.row_dimensions[fr].height = 28
    widths = [22, 18, *[20] * len(CROSS_ROUTE_COLS), 12, 42]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "B4"
    return ws2


def _import_summary(loaded: dict) -> str:
    return " · ".join(
        f"{k}:{len(loaded[k][0])}"
        for k in sorted(
            loaded, key=lambda x: FAMILY_ORDER.index(x) if x in FAMILY_ORDER else 99
        )
    )


def _route_index(sources: list, loaded: dict) -> str:
    return "；".join(
        f"{label}={path.name}({len(loaded[key][0])})"
        for path, label, key in sources
        if key in loaded
    )


def build():
    loaded_t, pairs_t, routes_t, cross_t = load_modality(SOURCES_TEXT, "text")
    loaded_i, pairs_i, routes_i, cross_i = load_modality(SOURCES_IMAGE, "image")
    loaded_v, pairs_v, routes_v, cross_v = load_modality(SOURCES_VIDEO, "video")

    wb = Workbook()
    wsr = wb.active
    wsr.title = SHEET_00
    for i, (a, b) in enumerate(
        [
            ("文件名", OUT.name),
            ("形态", "一本总册 · 8 Sheet；后台分册下载另议"),
            ("生成·商务", "commercial-billing/scripts/rebuild_discount_tier_workbook.py"),
            (
                "生成·解析/外发",
                "pricing/scripts/build_outward_quote_standard.py（回写 01 + 外发 xlsx）",
            ),
            ("SOP", "discount-tier-workbook-sop.md"),
            (SHEET_01, "报价依据：全量解析 / 原价专项 / 停用更低进价（外发脚本回写）"),
            (SHEET_10, "生文 · 成本族 × 对内阶梯（含GM）× 模型清单"),
            (SHEET_11, "生文 · 跨折同名 · 优先级/权重 · 推荐成本折"),
            (SHEET_20, "生图 · 成本族 × 对内阶梯 × 模型清单"),
            (SHEET_21, "生图 · 跨折同名 · 推荐成本折"),
            (SHEET_30, "生视频 · 成本族 × 对内阶梯 × 模型清单"),
            (SHEET_31, "生视频 · 跨折同名 · 推荐成本折"),
            (
                "外发文件",
                f"pricing/output/{OUTWARD_FILE}（01_生文/02_生图/03_生视频，整本可发）",
            ),
            (
                "档名",
                "对内五档：$1k/$5k/$10k/$30k/$50k（Plus…Enterprise）｜"
                "对外三档：$5k/$10k/$50k（表头「对外·」）",
            ),
            ("生文已导入", _import_summary(loaded_t)),
            ("生文交叉模型数", str(len(cross_t))),
            ("生图已导入", _import_summary(loaded_i)),
            ("生图交叉模型数", str(len(cross_i))),
            ("生视频已导入", _import_summary(loaded_v)),
            ("生视频交叉模型数", str(len(cross_v))),
            ("线路源·生文", f"pricing/input/{ROUTES_TEXT.name}/"),
            ("线路源索引·生文", _route_index(SOURCES_TEXT, loaded_t)),
            ("线路源·生图", f"pricing/input/{ROUTES_IMAGE.name}/"),
            ("线路源索引·生图", _route_index(SOURCES_IMAGE, loaded_i)),
            ("线路源·生视频", f"pricing/input/{ROUTES_VIDEO.name}/"),
            ("线路源索引·生视频", _route_index(SOURCES_VIDEO, loaded_v)),
            ("不含", "各折扣 src_* 整表；原料只归档 input"),
        ],
        1,
    ):
        wsr.cell(i, 1, a).font = font(bold=True)
        wsr.cell(i, 2, b).font = font()
        wsr.cell(i, 2).alignment = Alignment(wrap_text=True)
    wsr.column_dimensions["A"].width = 22
    wsr.column_dimensions["B"].width = 88

    ws01 = wb.create_sheet(SHEET_01, 1)
    ws01["A1"] = f"{SHEET_01} · 待回写"
    ws01["A1"].font = font(bold=True, size=12, color="B91C1C")
    ws01["A2"] = (
        "本页为对外报价依据（含启用/停用线路解析）。"
        "请在商务回灌后运行：python3 pricing/scripts/build_outward_quote_standard.py"
    )
    ws01["A2"].font = font(size=10, color="64748B")
    ws01["A2"].alignment = Alignment(wrap_text=True)
    ws01.row_dimensions[2].height = 36
    ws01.column_dimensions["A"].width = 96

    write_main_sheet(wb, SHEET_10, pairs_t, routes_t, cross_t, SHEET_11)
    write_cross_sheet(wb, SHEET_11, routes_t, cross_t, SHEET_10)
    write_main_sheet(wb, SHEET_20, pairs_i, routes_i, cross_i, SHEET_21)
    write_cross_sheet(wb, SHEET_21, routes_i, cross_i, SHEET_20)
    write_main_sheet(wb, SHEET_30, pairs_v, routes_v, cross_v, SHEET_31)
    write_cross_sheet(wb, SHEET_31, routes_v, cross_v, SHEET_30)

    order = [
        SHEET_00,
        SHEET_01,
        SHEET_10,
        SHEET_11,
        SHEET_20,
        SHEET_21,
        SHEET_30,
        SHEET_31,
    ]
    for idx, name in enumerate(order):
        wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    wb.save(OUT)
    patch_xlsx_for_wechat(OUT)
    print(f"saved {OUT}")
    print(f"sheets: {wb.sheetnames}")
    print(
        f"text cross={len(cross_t)} image cross={len(cross_i)} "
        f"video cross={len(cross_v)}"
    )


if __name__ == "__main__":
    build()
